#!/usr/bin/env python3
"""Idempotent catalog_mapping seed from a SKU-Square Mapping export.
Usage: gen-catalog-seed.py <mapping.xlsx> <out.sql> <client_id>
Loads rows with a Square Catalog ID (col D); maps sheet cols A-T. Status coerced to the
mapping_status enum (unknown/empty -> PENDING)."""
import openpyxl, sys, datetime
xlsx,out,client=sys.argv[1],sys.argv[2],sys.argv[3]
COLS=['source_row','vendor','vendor_sku','internal_ref','square_item_id','square_variation_id',
 'item_description','item_name','variation_name','times_ordered','status','tags','image_name',
 'image_url','wholesale_price','retail_price','first_seen','last_ordered','gems','notes','orientation']
IDX={'vendor':0,'vendor_sku':1,'internal_ref':2,'square_item_id':3,'square_variation_id':4,
 'item_description':5,'item_name':6,'variation_name':7,'times_ordered':8,'status':9,'tags':10,
 'image_name':11,'image_url':12,'wholesale_price':13,'retail_price':14,'first_seen':15,
 'last_ordered':16,'gems':17,'notes':18,'orientation':19}
ENUM={'PENDING','TAGGED','ENRICHED','NO_IMAGE','NEEDS_REVIEW','ACCESSORY','PUSHED'}
INT={'times_ordered'}; NUM={'wholesale_price','retail_price'}; DATE={'first_seen','last_ordered'}
def esc(v): return "'"+str(v).replace("'","''")+"'"
def val(c,v):
    if c=='status':
        if v in (None,''): return "'PENDING'"
        n=str(v).strip().upper().replace(' ','_').replace('-','_')
        return "'"+n+"'" if n in ENUM else "'PENDING'"
    if v in (None,''): return 'NULL'
    if c in INT:
        try: return str(int(float(v)))
        except: return 'NULL'
    if c in NUM:
        try: return format(float(v),'g')
        except: return 'NULL'
    if c in DATE:
        if isinstance(v,datetime.datetime): return "'"+v.date().isoformat()+"'"
        if isinstance(v,datetime.date): return "'"+v.isoformat()+"'"
    return esc(v)
wb=openpyxl.load_workbook(xlsx,read_only=True,data_only=True); ws=wb[wb.sheetnames[0]]
tups=[]; r=1
for row in ws.iter_rows(min_row=2,values_only=True):
    r+=1
    D=row[3] if len(row)>3 else None
    if D in (None,''): continue
    vs=[esc(client)]
    for c in COLS:
        if c=='source_row': vs.append(str(r)); continue
        i=IDX[c]; vs.append(val(c,row[i] if len(row)>i else None))
    tups.append("("+",".join(vs)+")")
with open(out,'w') as f:
    f.write("-- catalog_mapping seed for "+client+" ("+str(len(tups))+" rows). Idempotent.\n")
    f.write("begin;\ninsert into clients (id,name) values ("+esc(client)+","+esc(client)+") on conflict (id) do nothing;\n")
    f.write("delete from catalog_mapping where client_id = "+esc(client)+";\n")
    f.write("insert into catalog_mapping (client_id,"+",".join(COLS)+") values\n"+",\n".join(tups)+";\ncommit;\n")
print("wrote",len(tups),"rows ->",out)
